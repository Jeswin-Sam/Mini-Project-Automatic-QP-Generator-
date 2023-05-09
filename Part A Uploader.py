from firebase import firebase
import openpyxl


firebase = firebase.FirebaseApplication('https://automatic-qp-generator-7d88c-default-rtdb.firebaseio.com/', None)

current_sem = 'even'

dept = 'CSE'
year = '3rd year'
subject = 'Distributed Systems'
unit = 'Unit-1'

file_name = "Unit-1 Part A.xlsx"
database_path = f'/test/{dept}/{year}/{current_sem}/{subject}/question_bank/part_a/{unit}'


wb_obj = openpyxl.load_workbook(file_name)
sheet_obj = wb_obj.active

cell_obj = sheet_obj.cell(row = 1, column = 1)


for row in sheet_obj.iter_rows(values_only=True):
    data =  {'question': row[1],
             'bloom': row[2],
             'map': row[3]
            }
    result = firebase.post(database_path, data)
    
print(str(sheet_obj.max_row) + " questions uploaded successfully!")