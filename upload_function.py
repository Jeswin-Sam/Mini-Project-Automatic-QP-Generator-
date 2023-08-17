from firebase import firebase
import openpyxl

# ************************************Start of function definitions************************************     


# ------------------------for uploading part A------------------------
def upload_part_a(sheet_obj, database_path):
    """
    Uploads questions from the given sheet object to the specified database path for part A.

    Parameters:
        - sheet_obj (Worksheet object): The sheet object containing the questions.
        - database_path (str): The path to the database for storing the questions.

    Returns:
        None

    """

    # Iterate through each row in the sheet
    for row in sheet_obj.iter_rows(values_only=True):
        # Extract the question data from the row
        data = {
            'question': row[1],
            'bloom': row[2],
            'map': row[3]
        }
        
        # Upload the question data to the database
        result = firebase.post(database_path, data)

    # Print the number of questions uploaded successfully
    print(str(sheet_obj.max_row) + " questions uploaded successfully!")




# ------------------------for uploading part B------------------------
def upload_part_b(sheet_obj, database_path):
    for row in sheet_obj.iter_rows(values_only=True):
                data =  {'question': row[1].strip(),
                        'marks' : get_marks(row[1].strip()),
                        'bloom': row[2],
                        'map': row[3]
                        }
                result = firebase.post(database_path, data)
                
    print(str(sheet_obj.max_row) + " questions uploaded successfully!")
    


# ------------------------to get the marks for part B questions------------------------
def get_marks(question):
    """
    Extracts the marks from the given question string for part B questions.

    Parameters:
        - question (str): The question string.

    Returns:
        - marks (int or str): The extracted marks from the question. If marks are not found, returns "NA".

    """

    marks = ''  # Initialize an empty string to store the marks
    
    # Iterate through the question string in reverse order
    for i in range(len(question) - 2, -1, -1):
        if question[i] == '(':
            break
        else:
            marks += question[i]  # Add the character to the marks string
            
    marks = marks[::-1]  # Reverse the marks string
    
    try:
        marks = int(marks)  # Convert marks string to an integer
    except ValueError:
        marks = "NA"  # Set marks to "NA" if conversion fails
        
    return marks

        
        

def upload_questions(dept, year, subject, unit, part, file_name):
    """
    Uploads questions from an Excel file to a database based on the given parameters.

    Parameters:
        - dept (str): The department name.
        - year (str): The academic year.
        - subject (str): The subject name.
        - unit (str): The unit number.
        - part (str): The part of the question paper ('A' or 'B').
        - file_name (str): The path to the Excel file.

    Returns:
        None

    """

    current_sem = 'even'

    # Construct the database path based on the given parameters
    database_path = f'/test/{dept}/{year}/{current_sem}/{subject}/question_bank/part_{part}/Unit-{unit}'

    # Load the Excel file
    wb_obj = openpyxl.load_workbook(file_name)
    sheet_obj = wb_obj.active

    # Determine the part (Part A or Part B) and call the corresponding upload function
    if (part == 'A' or part == 'a'):
        upload_part_a(sheet_obj, database_path)
    else:
        upload_part_b(sheet_obj, database_path)

        

# ************************************End of function definitions************************************      
        
        
        
# Initialize firebase        
firebase = firebase.FirebaseApplication('https://automatic-qp-generator-7d88c-default-rtdb.firebaseio.com/', None)
