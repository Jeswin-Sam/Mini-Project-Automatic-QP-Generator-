from firebase import firebase
import openpyxl


firebase = firebase.FirebaseApplication('https://automatic-qp-generator-7d88c-default-rtdb.firebaseio.com/', None)


file_name = "Unit-1 Part B.xlsx"
database_path = '/test/part_b'


wb_obj = openpyxl.load_workbook(file_name)
sheet_obj = wb_obj.active

cell_obj = sheet_obj.cell(row = 1, column = 1)

def get_marks(question):

    marks = ''
    for i in range(len(question)-2,-1,-1):
        if question[i] == '(':
            break
        else:
            marks+=(question[i])
    marks = marks[::-1]
    
    try:
        marks = int(marks)
    except ValueError:
        marks = "NA"
        
    return marks


for row in sheet_obj.iter_rows(values_only=True):
    data =  {'question': row[1].strip(),
            'marks' : get_marks(row[1].strip()),
            'bloom': row[2],
            'map': row[3]
            }
    result = firebase.post(database_path, data)
    
print(str(sheet_obj.max_row) + " questions uploaded successfully!")